from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, File, UploadFile, WebSocket, WebSocketDisconnect
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse, FileResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Dict
from datetime import datetime, timedelta
from passlib.context import CryptContext
import jwt
import os
import logging
import json
import asyncio
from pathlib import Path
from bson import ObjectId
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Security
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
SECRET_KEY = os.getenv("SECRET_KEY", "your-secret-key-change-in-production")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30 * 24 * 60  # 30 days

security = HTTPBearer()

app = FastAPI()
api_router = APIRouter(prefix="/api")

# ==================== WebSocket Connection Manager ====================

class ConnectionManager:
    def __init__(self):
        self.active_connections: Dict[str, List[WebSocket]] = {}
        self.user_connections: Dict[str, WebSocket] = {}
    
    async def connect(self, websocket: WebSocket, user_id: str = None):
        await websocket.accept()
        if "all" not in self.active_connections:
            self.active_connections["all"] = []
        self.active_connections["all"].append(websocket)
        if user_id:
            self.user_connections[user_id] = websocket
    
    def disconnect(self, websocket: WebSocket, user_id: str = None):
        if "all" in self.active_connections:
            if websocket in self.active_connections["all"]:
                self.active_connections["all"].remove(websocket)
        if user_id and user_id in self.user_connections:
            del self.user_connections[user_id]
    
    async def broadcast(self, message: dict):
        """Broadcast message to all connected clients"""
        if "all" in self.active_connections:
            disconnected = []
            for connection in self.active_connections["all"]:
                try:
                    await connection.send_json(message)
                except:
                    disconnected.append(connection)
            for conn in disconnected:
                self.active_connections["all"].remove(conn)
    
    async def send_to_user(self, user_id: str, message: dict):
        """Send message to specific user"""
        if user_id in self.user_connections:
            try:
                await self.user_connections[user_id].send_json(message)
            except:
                del self.user_connections[user_id]

ws_manager = ConnectionManager()

# ==================== Models ====================

class UserRole:
    MANAGER = "manager"
    INSPECTOR = "inspector"
    CONTRACTOR = "contractor"
    AUTHORITY = "authority"

class SnagStatus:
    OPEN = "open"
    IN_PROGRESS = "in_progress"
    RESOLVED = "resolved"
    VERIFIED = "verified"

class SnagPriority:
    HIGH = "high"
    MEDIUM = "medium"
    LOW = "low"

# User Models
class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: str
    phone: Optional[str] = None

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserResponse(BaseModel):
    id: str
    email: str
    name: str
    role: str
    phone: Optional[str] = None
    push_token: Optional[str] = None

class TokenResponse(BaseModel):
    access_token: str
    token_type: str
    user: UserResponse

# Snag Models
class SnagCreate(BaseModel):
    description: str
    location: str
    project_name: str
    possible_solution: Optional[str] = None
    utm_coordinates: Optional[str] = None
    photos: List[str] = []  # base64 encoded images
    priority: str = SnagPriority.MEDIUM
    cost_estimate: Optional[float] = None
    assigned_contractor_id: Optional[str] = None
    assigned_authority_id: Optional[str] = None  # Deprecated, kept for backward compatibility
    assigned_authority_ids: List[str] = []  # New: multiple authorities
    due_date: Optional[datetime] = None

class SnagUpdate(BaseModel):
    description: Optional[str] = None
    location: Optional[str] = None
    project_name: Optional[str] = None
    possible_solution: Optional[str] = None
    utm_coordinates: Optional[str] = None
    photos: Optional[List[str]] = None
    status: Optional[str] = None
    priority: Optional[str] = None
    cost_estimate: Optional[float] = None
    assigned_contractor_id: Optional[str] = None
    assigned_authority_id: Optional[str] = None  # Deprecated, kept for backward compatibility
    assigned_authority_ids: Optional[List[str]] = None  # New: multiple authorities
    due_date: Optional[datetime] = None
    authority_feedback: Optional[str] = None
    authority_comment: Optional[str] = None
    work_started_date: Optional[datetime] = None
    work_completed_date: Optional[datetime] = None
    contractor_completion_date: Optional[datetime] = None
    contractor_completed: Optional[bool] = None
    authority_approved: Optional[bool] = None

class SnagResponse(BaseModel):
    id: str
    query_no: int
    description: str
    location: str
    project_name: str
    possible_solution: Optional[str]
    utm_coordinates: Optional[str]
    photos: List[str]
    status: str
    priority: str
    cost_estimate: Optional[float]
    assigned_contractor_id: Optional[str]
    assigned_contractor_name: Optional[str]
    assigned_authority_id: Optional[str]  # Deprecated, kept for backward compatibility
    assigned_authority_name: Optional[str]  # Deprecated, kept for backward compatibility
    assigned_authority_ids: List[str] = []  # New: multiple authority IDs
    assigned_authority_names: List[str] = []  # New: multiple authority names
    due_date: Optional[datetime]
    authority_feedback: Optional[str]
    authority_comment: Optional[str]
    created_by_id: str
    created_by_name: str
    created_at: datetime
    updated_at: datetime
    work_started_date: Optional[datetime]
    work_completed_date: Optional[datetime]
    contractor_completion_date: Optional[datetime]

# Notification Models
class NotificationCreate(BaseModel):
    user_id: str
    snag_id: str
    message: str

class NotificationResponse(BaseModel):
    id: str
    user_id: str
    snag_id: str
    message: str
    read: bool
    created_at: datetime

class PushTokenUpdate(BaseModel):
    push_token: str

# ==================== Helper Functions ====================

def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password):
    return pwd_context.hash(password)

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid authentication credentials")
        
        user = await db.users.find_one({"_id": ObjectId(user_id)})
        if user is None:
            raise HTTPException(status_code=401, detail="User not found")
        
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token has expired")
    except jwt.JWTError:
        raise HTTPException(status_code=401, detail="Could not validate credentials")
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid authentication credentials")

def require_role(allowed_roles: List[str]):
    async def role_checker(current_user: dict = Depends(get_current_user)):
        if current_user["role"] not in allowed_roles:
            raise HTTPException(status_code=403, detail="Not enough permissions")
        return current_user
    return role_checker

async def send_notification(user_id: str, snag_id: str, message: str):
    notification = {
        "user_id": user_id,
        "snag_id": snag_id,
        "message": message,
        "read": False,
        "created_at": datetime.utcnow()
    }
    await db.notifications.insert_one(notification)

async def get_next_query_no(project_name: str):
    """Get next query number for a specific project"""
    last_snag = await db.snags.find_one(
        {"project_name": project_name},
        sort=[("query_no", -1)]
    )
    return (last_snag["query_no"] + 1) if last_snag else 1

# ==================== Auth Endpoints ====================

@api_router.post("/auth/register", response_model=UserResponse)
async def register(
    user_data: UserCreate,
    current_user: dict = Depends(require_role([UserRole.MANAGER]))
):
    # Check if user exists
    existing_user = await db.users.find_one({"email": user_data.email})
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Validate role
    valid_roles = [UserRole.MANAGER, UserRole.INSPECTOR, UserRole.CONTRACTOR, UserRole.AUTHORITY]
    if user_data.role not in valid_roles:
        raise HTTPException(status_code=400, detail="Invalid role")
    
    # Create user
    user_dict = {
        "email": user_data.email,
        "password": get_password_hash(user_data.password),
        "name": user_data.name,
        "role": user_data.role,
        "phone": user_data.phone,
        "push_token": None,
        "created_at": datetime.utcnow()
    }
    
    result = await db.users.insert_one(user_dict)
    
    return UserResponse(
        id=str(result.inserted_id),
        email=user_data.email,
        name=user_data.name,
        role=user_data.role,
        phone=user_data.phone
    )

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(user_data: UserLogin):
    user = await db.users.find_one({"email": user_data.email})
    if not user or not verify_password(user_data.password, user["password"]):
        raise HTTPException(status_code=401, detail="Incorrect email or password")
    
    access_token = create_access_token(data={"sub": str(user["_id"])})
    
    return TokenResponse(
        access_token=access_token,
        token_type="bearer",
        user=UserResponse(
            id=str(user["_id"]),
            email=user["email"],
            name=user["name"],
            role=user["role"],
            phone=user.get("phone"),
            push_token=user.get("push_token")
        )
    )

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(current_user: dict = Depends(get_current_user)):
    return UserResponse(
        id=str(current_user["_id"]),
        email=current_user["email"],
        name=current_user["name"],
        role=current_user["role"],
        phone=current_user.get("phone"),
        push_token=current_user.get("push_token")
    )

@api_router.put("/auth/push-token")
async def update_push_token(
    token_data: PushTokenUpdate,
    current_user: dict = Depends(get_current_user)
):
    await db.users.update_one(
        {"_id": current_user["_id"]},
        {"$set": {"push_token": token_data.push_token}}
    )
    return {"message": "Push token updated successfully"}

# ==================== User Management ====================

@api_router.get("/users", response_model=List[UserResponse])
async def get_users(current_user: dict = Depends(get_current_user)):
    users = await db.users.find().to_list(1000)
    return [
        UserResponse(
            id=str(user["_id"]),
            email=user["email"],
            name=user["name"],
            role=user["role"],
            phone=user.get("phone"),
            push_token=user.get("push_token")
        )
        for user in users
    ]

@api_router.get("/users/contractors", response_model=List[UserResponse])
async def get_contractors(current_user: dict = Depends(get_current_user)):
    contractors = await db.users.find({"role": UserRole.CONTRACTOR}).to_list(1000)
    return [
        UserResponse(
            id=str(user["_id"]),
            email=user["email"],
            name=user["name"],
            role=user["role"],
            phone=user.get("phone")
        )
        for user in contractors
    ]

@api_router.get("/users/authorities", response_model=List[UserResponse])
async def get_authorities(current_user: dict = Depends(get_current_user)):
    authorities = await db.users.find({"role": UserRole.AUTHORITY}).to_list(1000)
    return [
        UserResponse(
            id=str(user["_id"]),
            email=user["email"],
            name=user["name"],
            role=user["role"],
            phone=user.get("phone")
        )
        for user in authorities
    ]

@api_router.get("/buildings/{building_name}/suggested-authorities")
async def get_suggested_authorities_for_building(
    building_name: str,
    current_user: dict = Depends(get_current_user)
):
    """Get suggested authorities based on historical snag data for this building"""
    # Aggregate to find most common authorities for this building
    pipeline = [
        {
            "$match": {
                "project_name": building_name,
                "assigned_authority_id": {"$exists": True, "$ne": None}
            }
        },
        {
            "$group": {
                "_id": "$assigned_authority_id",
                "snag_count": {"$sum": 1},
                "last_assigned": {"$max": "$created_at"}
            }
        },
        {
            "$sort": {"snag_count": -1, "last_assigned": -1}
        },
        {
            "$limit": 3
        }
    ]
    
    authority_stats = await db.snags.aggregate(pipeline).to_list(10)
    
    suggested = []
    for stat in authority_stats:
        try:
            authority = await db.users.find_one({"_id": ObjectId(stat["_id"])})
            if authority:
                suggested.append({
                    "id": str(authority["_id"]),
                    "name": authority["name"],
                    "snag_count": stat["snag_count"]
                })
        except:
            continue
    
    return {"suggested_authorities": suggested}

@api_router.get("/buildings/{building_name}/previous-authority")
async def get_previous_authority_for_building(
    building_name: str,
    current_user: dict = Depends(get_current_user)
):
    """Get the authority assigned to the most recent snag for this building"""
    # Find the most recent snag for this building that has an assigned authority
    last_snag = await db.snags.find_one(
        {
            "project_name": building_name,
            "assigned_authority_id": {"$exists": True, "$ne": None}
        },
        sort=[("created_at", -1)]
    )
    
    if last_snag and last_snag.get("assigned_authority_id"):
        authority = await db.users.find_one({"_id": ObjectId(last_snag["assigned_authority_id"])})
        if authority:
            return {
                "authority_id": str(authority["_id"]),
                "authority_name": authority["name"]
            }
    
    return {"authority_id": None, "authority_name": None}

# ==================== Snag Endpoints ====================

@api_router.post("/snags", response_model=SnagResponse)
async def create_snag(
    snag_data: SnagCreate,
    current_user: dict = Depends(require_role([UserRole.MANAGER, UserRole.INSPECTOR]))
):
    # Get project-specific query number
    query_no = await get_next_query_no(snag_data.project_name)
    
    # Auto-assign authority from previous snag for this building if not provided
    # Handle multiple authorities - combine old single ID with new array for backward compatibility
    assigned_authority_ids = list(snag_data.assigned_authority_ids) if snag_data.assigned_authority_ids else []
    
    # If old single authority ID is provided and not in array, add it
    if snag_data.assigned_authority_id and snag_data.assigned_authority_id not in assigned_authority_ids:
        assigned_authority_ids.append(snag_data.assigned_authority_id)
    
    # If no authorities assigned, try to get from last snag for this building
    if not assigned_authority_ids:
        last_snag = await db.snags.find_one(
            {
                "project_name": snag_data.project_name,
                "$or": [
                    {"assigned_authority_ids": {"$exists": True, "$ne": []}},
                    {"assigned_authority_id": {"$exists": True, "$ne": None}}
                ]
            },
            sort=[("created_at", -1)]
        )
        if last_snag:
            if last_snag.get("assigned_authority_ids"):
                assigned_authority_ids = last_snag["assigned_authority_ids"]
            elif last_snag.get("assigned_authority_id"):
                assigned_authority_ids = [last_snag["assigned_authority_id"]]
    
    # For backward compatibility, set single authority ID to first in list
    assigned_authority_id = assigned_authority_ids[0] if assigned_authority_ids else None
    
    snag_dict = {
        "query_no": query_no,
        "description": snag_data.description,
        "location": snag_data.location,
        "project_name": snag_data.project_name,
        "possible_solution": snag_data.possible_solution,
        "utm_coordinates": snag_data.utm_coordinates,
        "photos": snag_data.photos,
        "status": SnagStatus.OPEN,
        "priority": snag_data.priority,
        "cost_estimate": snag_data.cost_estimate,
        "assigned_contractor_id": snag_data.assigned_contractor_id,
        "assigned_authority_id": assigned_authority_id,  # Backward compatibility
        "assigned_authority_ids": assigned_authority_ids,  # New: multiple authorities
        "due_date": snag_data.due_date,
        "authority_feedback": None,
        "authority_comment": None,
        "created_by_id": str(current_user["_id"]),
        "created_by_name": current_user["name"],
        "created_at": datetime.utcnow(),
        "updated_at": datetime.utcnow(),
        "work_started_date": None,
        "work_completed_date": None,
        "contractor_completion_date": None,
        "contractor_completed": False,
        "authority_approved": False
    }
    
    result = await db.snags.insert_one(snag_dict)
    snag_id = str(result.inserted_id)
    
    # Send notifications to all assigned users
    notification_recipients = []
    
    # Send notification to assigned contractor
    if snag_data.assigned_contractor_id:
        notification_recipients.append(snag_data.assigned_contractor_id)
        await send_notification(
            snag_data.assigned_contractor_id,
            snag_id,
            f"New snag #{query_no} assigned to you at {snag_data.project_name} - {snag_data.location}"
        )
        await broadcast_notification(snag_data.assigned_contractor_id, {
            "snag_id": snag_id,
            "message": f"New snag #{query_no} assigned to you at {snag_data.project_name} - {snag_data.location}"
        })
    
    # Send notification to all assigned authorities
    for auth_id in assigned_authority_ids:
        if auth_id not in notification_recipients:
            notification_recipients.append(auth_id)
            await send_notification(
                auth_id,
                snag_id,
                f"New snag #{query_no} created at {snag_data.project_name} - {snag_data.location} (You are an assigned authority)"
            )
            await broadcast_notification(auth_id, {
                "snag_id": snag_id,
                "message": f"New snag #{query_no} created at {snag_data.project_name} - You are an assigned authority"
            })
    
    # Get contractor name
    contractor_name = None
    if snag_data.assigned_contractor_id:
        contractor = await db.users.find_one({"_id": ObjectId(snag_data.assigned_contractor_id)})
        contractor_name = contractor["name"] if contractor else None
    
    # Get all authority names
    authority_names = []
    for auth_id in assigned_authority_ids:
        try:
            authority = await db.users.find_one({"_id": ObjectId(auth_id)})
            if authority:
                authority_names.append(authority["name"])
        except:
            pass
    
    # Backward compatibility: first authority name
    authority_name = authority_names[0] if authority_names else None
    
    snag_response = SnagResponse(
        id=snag_id,
        query_no=query_no,
        description=snag_data.description,
        location=snag_data.location,
        project_name=snag_data.project_name,
        possible_solution=snag_data.possible_solution,
        utm_coordinates=snag_data.utm_coordinates,
        photos=snag_data.photos,
        status=SnagStatus.OPEN,
        priority=snag_data.priority,
        cost_estimate=snag_data.cost_estimate,
        assigned_contractor_id=snag_data.assigned_contractor_id,
        assigned_contractor_name=contractor_name,
        assigned_authority_id=assigned_authority_id,
        assigned_authority_name=authority_name,
        assigned_authority_ids=assigned_authority_ids,
        assigned_authority_names=authority_names,
        due_date=snag_data.due_date,
        authority_feedback=None,
        authority_comment=None,
        created_by_id=str(current_user["_id"]),
        created_by_name=current_user["name"],
        created_at=snag_dict["created_at"],
        updated_at=snag_dict["updated_at"],
        work_started_date=None,
        work_completed_date=None,
        contractor_completion_date=None
    )
    
    # Broadcast snag creation to all connected clients
    await broadcast_snag_update("created", snag_response.dict())
    
    return snag_response

@api_router.get("/snags", response_model=List[SnagResponse])
async def get_snags(
    status: Optional[str] = None,
    priority: Optional[str] = None,
    location: Optional[str] = None,
    project_name: Optional[str] = None,
    assigned_contractor_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    
    # Role-based filtering
    if current_user["role"] == UserRole.CONTRACTOR:
        query["assigned_contractor_id"] = str(current_user["_id"])
    elif current_user["role"] == UserRole.AUTHORITY:
        # Authority can see snags where they are in either old or new field
        user_id = str(current_user["_id"])
        query["$or"] = [
            {"assigned_authority_id": user_id},
            {"assigned_authority_ids": user_id}
        ]
    
    if status:
        query["status"] = status
    if priority:
        query["priority"] = priority
    if location:
        query["location"] = {"$regex": location, "$options": "i"}
    if project_name:
        query["project_name"] = {"$regex": project_name, "$options": "i"}
    if assigned_contractor_id:
        query["assigned_contractor_id"] = assigned_contractor_id
    
    snags = await db.snags.find(query).sort("created_at", -1).to_list(1000)
    
    # Get contractor and authority names
    result = []
    for snag in snags:
        contractor_name = None
        if snag.get("assigned_contractor_id"):
            contractor = await db.users.find_one({"_id": ObjectId(snag["assigned_contractor_id"])})
            contractor_name = contractor["name"] if contractor else None
        
        # Get multiple authority names
        assigned_authority_ids = snag.get("assigned_authority_ids", [])
        # Backward compatibility: if old field exists and not in new array
        if snag.get("assigned_authority_id") and snag["assigned_authority_id"] not in assigned_authority_ids:
            assigned_authority_ids = [snag["assigned_authority_id"]] + assigned_authority_ids
        
        authority_names = []
        for auth_id in assigned_authority_ids:
            try:
                authority = await db.users.find_one({"_id": ObjectId(auth_id)})
                if authority:
                    authority_names.append(authority["name"])
            except:
                pass
        
        # Backward compatibility
        authority_name = authority_names[0] if authority_names else None
        assigned_authority_id = assigned_authority_ids[0] if assigned_authority_ids else snag.get("assigned_authority_id")
        
        result.append(SnagResponse(
            id=str(snag["_id"]),
            query_no=snag["query_no"],
            description=snag["description"],
            location=snag["location"],
            project_name=snag.get("project_name", ""),
            possible_solution=snag.get("possible_solution"),
            utm_coordinates=snag.get("utm_coordinates"),
            photos=snag["photos"],
            status=snag["status"],
            priority=snag["priority"],
            cost_estimate=snag.get("cost_estimate"),
            assigned_contractor_id=snag.get("assigned_contractor_id"),
            assigned_contractor_name=contractor_name,
            assigned_authority_id=assigned_authority_id,
            assigned_authority_name=authority_name,
            assigned_authority_ids=assigned_authority_ids,
            assigned_authority_names=authority_names,
            due_date=snag.get("due_date"),
            authority_feedback=snag.get("authority_feedback"),
            authority_comment=snag.get("authority_comment"),
            created_by_id=snag["created_by_id"],
            created_by_name=snag["created_by_name"],
            created_at=snag["created_at"],
            updated_at=snag["updated_at"],
            work_started_date=snag.get("work_started_date"),
            work_completed_date=snag.get("work_completed_date"),
            contractor_completion_date=snag.get("contractor_completion_date")
        ))
    
    return result

@api_router.get("/snags/{snag_id}", response_model=SnagResponse)
async def get_snag(
    snag_id: str,
    current_user: dict = Depends(get_current_user)
):
    snag = await db.snags.find_one({"_id": ObjectId(snag_id)})
    if not snag:
        raise HTTPException(status_code=404, detail="Snag not found")
    
    contractor_name = None
    if snag.get("assigned_contractor_id"):
        contractor = await db.users.find_one({"_id": ObjectId(snag["assigned_contractor_id"])})
        contractor_name = contractor["name"] if contractor else None
    
    authority_name = None
    if snag.get("assigned_authority_id"):
        authority = await db.users.find_one({"_id": ObjectId(snag["assigned_authority_id"])})
        authority_name = authority["name"] if authority else None
    
    return SnagResponse(
        id=str(snag["_id"]),
        query_no=snag["query_no"],
        description=snag["description"],
        location=snag["location"],
        project_name=snag.get("project_name", ""),
        possible_solution=snag.get("possible_solution"),
        utm_coordinates=snag.get("utm_coordinates"),
        photos=snag["photos"],
        status=snag["status"],
        priority=snag["priority"],
        cost_estimate=snag.get("cost_estimate"),
        assigned_contractor_id=snag.get("assigned_contractor_id"),
        assigned_contractor_name=contractor_name,
        assigned_authority_id=snag.get("assigned_authority_id"),
        assigned_authority_name=authority_name,
        due_date=snag.get("due_date"),
        authority_feedback=snag.get("authority_feedback"),
        authority_comment=snag.get("authority_comment"),
        created_by_id=snag["created_by_id"],
        created_by_name=snag["created_by_name"],
        created_at=snag["created_at"],
        updated_at=snag["updated_at"],
        work_started_date=snag.get("work_started_date"),
        work_completed_date=snag.get("work_completed_date"),
        contractor_completion_date=snag.get("contractor_completion_date")
    )

@api_router.put("/snags/{snag_id}", response_model=SnagResponse)
async def update_snag(
    snag_id: str,
    snag_update: SnagUpdate,
    current_user: dict = Depends(get_current_user)
):
    snag = await db.snags.find_one({"_id": ObjectId(snag_id)})
    if not snag:
        raise HTTPException(status_code=404, detail="Snag not found")
    
    # Build update dict
    update_data = {k: v for k, v in snag_update.dict(exclude_unset=True).items() if v is not None}
    update_data["updated_at"] = datetime.utcnow()
    
    # Role-based permissions
    if current_user["role"] == UserRole.CONTRACTOR:
        if snag.get("assigned_contractor_id") != str(current_user["_id"]):
            raise HTTPException(status_code=403, detail="Not assigned to this snag")
        
        # Contractors can mark completion AND set completion date
        allowed_updates = {"contractor_completed", "work_started_date", "work_completed_date", "contractor_completion_date"}
        provided_updates = {k for k, v in snag_update.dict(exclude_unset=True).items() if v is not None}
        if not provided_updates.issubset(allowed_updates):
            raise HTTPException(status_code=403, detail="Contractors can only mark completion and set completion date")
        
        # When contractor marks as completed
        if update_data.get("contractor_completed"):
            update_data["contractor_completed"] = True
            update_data["work_completed_date"] = datetime.utcnow()
            if snag["status"] == SnagStatus.OPEN:
                update_data["status"] = SnagStatus.IN_PROGRESS
            
            # Check if authority already approved
            if snag.get("authority_approved"):
                update_data["status"] = SnagStatus.RESOLVED
    
    elif current_user["role"] == UserRole.AUTHORITY:
        # Authority can approve, provide feedback AND add comments
        allowed_updates = {"authority_approved", "authority_feedback", "authority_comment", "status"}
        provided_updates = {k for k, v in snag_update.dict(exclude_unset=True).items() if v is not None}
        if not provided_updates.issubset(allowed_updates):
            raise HTTPException(status_code=403, detail="Authority can only approve, provide feedback and comments")
        
        # When authority approves
        if update_data.get("authority_approved"):
            update_data["authority_approved"] = True
            
            # Check if contractor already completed
            if snag.get("contractor_completed"):
                update_data["status"] = SnagStatus.RESOLVED
            else:
                if snag["status"] == SnagStatus.OPEN:
                    update_data["status"] = SnagStatus.IN_PROGRESS
    
    elif current_user["role"] in [UserRole.MANAGER, UserRole.INSPECTOR]:
        # Manager and Inspector can edit EVERYTHING
        # No restrictions on fields
        
        # Handle approval like authority if they set it
        if update_data.get("authority_approved"):
            update_data["authority_approved"] = True
            if snag.get("contractor_completed"):
                update_data["status"] = SnagStatus.RESOLVED
    
    old_status = snag["status"]
    new_status = update_data.get("status", old_status)
    
    await db.snags.update_one(
        {"_id": ObjectId(snag_id)},
        {"$set": update_data}
    )
    
    # Send notifications
    notifications_sent = []
    
    # Notify on contractor completion
    if update_data.get("contractor_completed") and not snag.get("contractor_completed"):
        # Notify authority and manager
        authorities = await db.users.find({"role": {"$in": [UserRole.AUTHORITY, UserRole.MANAGER]}}).to_list(100)
        for auth in authorities:
            await send_notification(
                str(auth["_id"]),
                snag_id,
                f"Snag #{snag['query_no']} completed by contractor - pending your approval"
            )
        notifications_sent.append("authority")
    
    # Notify on authority approval
    if update_data.get("authority_approved") and not snag.get("authority_approved"):
        # Notify contractor and creator
        if snag.get("assigned_contractor_id"):
            await send_notification(
                snag["assigned_contractor_id"],
                snag_id,
                f"Snag #{snag['query_no']} approved by authority"
            )
        await send_notification(
            snag["created_by_id"],
            snag_id,
            f"Snag #{snag['query_no']} approved by authority"
        )
        notifications_sent.append("contractor")
    
    # Notify on status change to resolved
    if old_status != new_status and new_status == SnagStatus.RESOLVED:
        if "authority" not in notifications_sent:
            await send_notification(
                snag["created_by_id"],
                snag_id,
                f"Snag #{snag['query_no']} marked as RESOLVED (Contractor completed & Authority approved)"
            )
        if snag.get("assigned_contractor_id") and "contractor" not in notifications_sent:
            await send_notification(
                snag["assigned_contractor_id"],
                snag_id,
                f"Snag #{snag['query_no']} marked as RESOLVED"
            )
    
    # Get updated snag
    updated_snag = await db.snags.find_one({"_id": ObjectId(snag_id)})
    
    contractor_name = None
    if updated_snag.get("assigned_contractor_id"):
        contractor = await db.users.find_one({"_id": ObjectId(updated_snag["assigned_contractor_id"])})
        contractor_name = contractor["name"] if contractor else None
    
    authority_name = None
    if updated_snag.get("assigned_authority_id"):
        authority = await db.users.find_one({"_id": ObjectId(updated_snag["assigned_authority_id"])})
        authority_name = authority["name"] if authority else None
    
    snag_response = SnagResponse(
        id=str(updated_snag["_id"]),
        query_no=updated_snag["query_no"],
        description=updated_snag["description"],
        location=updated_snag["location"],
        project_name=updated_snag.get("project_name", ""),
        possible_solution=updated_snag.get("possible_solution"),
        utm_coordinates=updated_snag.get("utm_coordinates"),
        photos=updated_snag["photos"],
        status=updated_snag["status"],
        priority=updated_snag["priority"],
        cost_estimate=updated_snag.get("cost_estimate"),
        assigned_contractor_id=updated_snag.get("assigned_contractor_id"),
        assigned_contractor_name=contractor_name,
        assigned_authority_id=updated_snag.get("assigned_authority_id"),
        assigned_authority_name=authority_name,
        due_date=updated_snag.get("due_date"),
        authority_feedback=updated_snag.get("authority_feedback"),
        authority_comment=updated_snag.get("authority_comment"),
        created_by_id=updated_snag["created_by_id"],
        created_by_name=updated_snag["created_by_name"],
        created_at=updated_snag["created_at"],
        updated_at=updated_snag["updated_at"],
        work_started_date=updated_snag.get("work_started_date"),
        work_completed_date=updated_snag.get("work_completed_date"),
        contractor_completion_date=updated_snag.get("contractor_completion_date")
    )
    
    # Broadcast snag update to all connected clients
    await broadcast_snag_update("updated", snag_response.dict())
    
    return snag_response

@api_router.delete("/snags/{snag_id}")
async def delete_snag(
    snag_id: str,
    current_user: dict = Depends(require_role([UserRole.MANAGER]))
):
    # Get snag info before deletion for broadcast
    snag = await db.snags.find_one({"_id": ObjectId(snag_id)})
    
    result = await db.snags.delete_one({"_id": ObjectId(snag_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Snag not found")
    
    # Broadcast deletion to all connected clients
    await broadcast_snag_update("deleted", {"id": snag_id, "query_no": snag.get("query_no") if snag else None})
    
    return {"message": "Snag deleted successfully"}

# ==================== Notification Endpoints ====================

@api_router.get("/notifications", response_model=List[NotificationResponse])
async def get_notifications(current_user: dict = Depends(get_current_user)):
    notifications = await db.notifications.find(
        {"user_id": str(current_user["_id"])}
    ).sort("created_at", -1).to_list(100)
    
    return [
        NotificationResponse(
            id=str(notif["_id"]),
            user_id=notif["user_id"],
            snag_id=notif["snag_id"],
            message=notif["message"],
            read=notif["read"],
            created_at=notif["created_at"]
        )
        for notif in notifications
    ]

@api_router.put("/notifications/{notification_id}/read")
async def mark_notification_read(
    notification_id: str,
    current_user: dict = Depends(get_current_user)
):
    await db.notifications.update_one(
        {"_id": ObjectId(notification_id), "user_id": str(current_user["_id"])},
        {"$set": {"read": True}}
    )
    return {"message": "Notification marked as read"}

@api_router.put("/notifications/read-all")
async def mark_all_notifications_read(current_user: dict = Depends(get_current_user)):
    await db.notifications.update_many(
        {"user_id": str(current_user["_id"]), "read": False},
        {"$set": {"read": True}}
    )
    return {"message": "All notifications marked as read"}

# ==================== Export Endpoints ====================

@api_router.get("/snags/export/excel")
async def export_snags_excel(
    status: Optional[str] = None,
    project_name: Optional[str] = None,
    current_user: dict = Depends(require_role([UserRole.MANAGER, UserRole.AUTHORITY]))
):
    query = {}
    if status:
        query["status"] = status
    if project_name:
        query["project_name"] = project_name
    
    snags = await db.snags.find(query).sort([("project_name", 1), ("query_no", 1)]).to_list(1000)
    
    # Group snags by project
    snags_by_project = {}
    for snag in snags:
        project = snag.get("project_name", "Uncategorized")
        if project not in snags_by_project:
            snags_by_project[project] = []
        snags_by_project[project].append(snag)
    
    # Create workbook
    wb = Workbook()
    
    # Create a sheet for each project
    for idx, (project, project_snags) in enumerate(snags_by_project.items()):
        if idx == 0:
            ws = wb.active
            ws.title = project[:31]  # Excel sheet name limit
        else:
            ws = wb.create_sheet(title=project[:31])
        
        # Headers
        headers = [
            "Query No", "Project/Building", "Location", "Description", "Possible Solution",
            "Status", "Priority", "Cost Estimate", "Assigned Contractor", "Due Date",
            "Created By", "Created Date", "UTM Coordinates", "Authority Feedback",
            "Work Started", "Work Completed", "Photo Count"
        ]
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        for row, snag in enumerate(project_snags, 2):
            contractor_name = ""
            if snag.get("assigned_contractor_id"):
                contractor = await db.users.find_one({"_id": ObjectId(snag["assigned_contractor_id"])})
                contractor_name = contractor["name"] if contractor else ""
            
            ws.cell(row=row, column=1, value=snag["query_no"])
            ws.cell(row=row, column=2, value=snag.get("project_name", ""))
            ws.cell(row=row, column=3, value=snag["location"])
            ws.cell(row=row, column=4, value=snag["description"])
            ws.cell(row=row, column=5, value=snag.get("possible_solution", ""))
            ws.cell(row=row, column=6, value=snag["status"])
            ws.cell(row=row, column=7, value=snag["priority"])
            ws.cell(row=row, column=8, value=snag.get("cost_estimate", ""))
            ws.cell(row=row, column=9, value=contractor_name)
            ws.cell(row=row, column=10, value=snag.get("due_date").strftime("%Y-%m-%d") if snag.get("due_date") else "")
            ws.cell(row=row, column=11, value=snag["created_by_name"])
            ws.cell(row=row, column=12, value=snag["created_at"].strftime("%Y-%m-%d %H:%M"))
            ws.cell(row=row, column=13, value=snag.get("utm_coordinates", ""))
            ws.cell(row=row, column=14, value=snag.get("authority_feedback", ""))
            ws.cell(row=row, column=15, value=snag.get("work_started_date").strftime("%Y-%m-%d %H:%M") if snag.get("work_started_date") else "")
            ws.cell(row=row, column=16, value=snag.get("work_completed_date").strftime("%Y-%m-%d %H:%M") if snag.get("work_completed_date") else "")
            ws.cell(row=row, column=17, value=len(snag.get("photos", [])))
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    # Save to bytes
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=snag_list_by_project.xlsx"}
    )

@api_router.get("/snags/export/pdf")
async def export_snags_pdf(
    status: Optional[str] = None,
    project_name: Optional[str] = None,
    current_user: dict = Depends(require_role([UserRole.MANAGER, UserRole.AUTHORITY]))
):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER
    import base64
    
    query = {}
    if status:
        query["status"] = status
    if project_name:
        query["project_name"] = project_name
    
    snags = await db.snags.find(query).sort([("project_name", 1), ("query_no", 1)]).to_list(1000)
    
    # Group snags by project
    snags_by_project = {}
    for snag in snags:
        project = snag.get("project_name", "Uncategorized")
        if project not in snags_by_project:
            snags_by_project[project] = []
        snags_by_project[project].append(snag)
    
    # Create PDF
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        pdf_buffer,
        pagesize=landscape(A4),
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.75*inch,
        bottomMargin=0.5*inch
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#366092'),
        spaceAfter=10
    )
    normal_style = styles['Normal']
    
    story = []
    
    # Title page
    story.append(Paragraph("PMC Snag List Report", title_style))
    story.append(Spacer(1, 0.3*inch))
    story.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", normal_style))
    story.append(Paragraph(f"Generated by: {current_user['name']}", normal_style))
    story.append(PageBreak())
    
    # Process each project
    for project, project_snags in snags_by_project.items():
        # Project header
        story.append(Paragraph(f"Project: {project}", heading_style))
        story.append(Paragraph(f"Total Snags: {len(project_snags)}", normal_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Each snag on new page
        for snag in project_snags:
            # Get contractor name
            contractor_name = "Not Assigned"
            if snag.get("assigned_contractor_id"):
                contractor = await db.users.find_one({"_id": ObjectId(snag["assigned_contractor_id"])})
                contractor_name = contractor["name"] if contractor else "Not Assigned"
            
            # Snag header
            story.append(Paragraph(f"Snag #{snag['query_no']}", heading_style))
            
            # Snag details table
            details_data = [
                ["Field", "Value"],
                ["Project/Building", snag.get("project_name", "")],
                ["Location", snag["location"]],
                ["Status", snag["status"].replace("_", " ").title()],
                ["Priority", snag["priority"].title()],
                ["Description", snag["description"]],
            ]
            
            if snag.get("possible_solution"):
                details_data.append(["Possible Solution", snag.get("possible_solution")])
            
            if snag.get("utm_coordinates"):
                details_data.append(["GPS Coordinates (UTM)", snag.get("utm_coordinates")])
            
            details_data.extend([
                ["Cost Estimate", f"${snag['cost_estimate']}" if snag.get("cost_estimate") else "Not Specified"],
                ["Assigned Contractor", contractor_name],
                ["Due Date", snag.get("due_date").strftime("%Y-%m-%d") if snag.get("due_date") else "Not Set"],
                ["Created By", snag["created_by_name"]],
                ["Created Date", snag["created_at"].strftime("%Y-%m-%d %H:%M")],
            ])
            
            if snag.get("work_started_date"):
                details_data.append(["Work Started", snag["work_started_date"].strftime("%Y-%m-%d %H:%M")])
            
            if snag.get("work_completed_date"):
                details_data.append(["Work Completed", snag["work_completed_date"].strftime("%Y-%m-%d %H:%M")])
            
            if snag.get("authority_feedback"):
                details_data.append(["Authority Feedback", snag["authority_feedback"]])
            
            details_table = Table(details_data, colWidths=[2.5*inch, 7*inch])
            details_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#e8f4f8')),
                ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            
            story.append(details_table)
            story.append(Spacer(1, 0.2*inch))
            
            # Photos
            if snag.get("photos"):
                story.append(Paragraph(f"Photos ({len(snag['photos'])})", heading_style))
                story.append(Spacer(1, 0.1*inch))
                
                # Display photos in rows of 3
                photos_per_row = 3
                photo_size = 2.5 * inch
                
                for i in range(0, len(snag["photos"]), photos_per_row):
                    photo_row = []
                    for j in range(photos_per_row):
                        if i + j < len(snag["photos"]):
                            photo_data = snag["photos"][i + j]
                            try:
                                # Extract base64 data
                                if ',' in photo_data:
                                    photo_data = photo_data.split(',', 1)[1]
                                
                                # Create image from base64
                                img_data = base64.b64decode(photo_data)
                                img_buffer = io.BytesIO(img_data)
                                
                                # Create ReportLab image
                                img = RLImage(img_buffer, width=photo_size, height=photo_size)
                                photo_row.append(img)
                            except Exception as e:
                                logger.error(f"Error processing photo: {e}")
                                photo_row.append(Paragraph("Photo Error", normal_style))
                        else:
                            photo_row.append("")
                    
                    if photo_row:
                        photo_table = Table([photo_row], colWidths=[photo_size + 0.2*inch] * photos_per_row)
                        photo_table.setStyle(TableStyle([
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ]))
                        story.append(photo_table)
                        story.append(Spacer(1, 0.1*inch))
            
            # New page for next snag
            story.append(PageBreak())
    
    # Build PDF
    doc.build(story)
    pdf_buffer.seek(0)
    
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=snag_list_by_project.pdf"}
    )

# ==================== Dashboard Stats ====================

@api_router.get("/projects/names")
async def get_project_names(current_user: dict = Depends(get_current_user)):
    """Get list of unique project names"""
    projects = await db.snags.distinct("project_name")
    return {"projects": sorted([p for p in projects if p])}

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(current_user: dict = Depends(get_current_user)):
    query = {}
    if current_user["role"] == UserRole.CONTRACTOR:
        query["assigned_contractor_id"] = str(current_user["_id"])
    
    total_snags = await db.snags.count_documents(query)
    open_snags = await db.snags.count_documents({**query, "status": SnagStatus.OPEN})
    in_progress_snags = await db.snags.count_documents({**query, "status": SnagStatus.IN_PROGRESS})
    resolved_snags = await db.snags.count_documents({**query, "status": SnagStatus.RESOLVED})
    verified_snags = await db.snags.count_documents({**query, "status": SnagStatus.VERIFIED})
    
    high_priority = await db.snags.count_documents({**query, "priority": SnagPriority.HIGH})
    
    return {
        "total_snags": total_snags,
        "open_snags": open_snags,
        "in_progress_snags": in_progress_snags,
        "resolved_snags": resolved_snags,
        "verified_snags": verified_snags,
        "high_priority": high_priority
    }

# ==================== WebSocket Endpoint ====================

@app.websocket("/api/ws")
async def websocket_endpoint(websocket: WebSocket):
    """WebSocket endpoint for real-time updates"""
    user_id = None
    try:
        # Accept connection first
        await ws_manager.connect(websocket)
        
        while True:
            try:
                # Wait for messages from client
                data = await websocket.receive_text()
                message = json.loads(data)
                
                # Handle authentication message
                if message.get("type") == "auth":
                    token = message.get("token")
                    if token:
                        try:
                            payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
                            user_id = payload.get("sub")
                            if user_id:
                                ws_manager.user_connections[user_id] = websocket
                                await websocket.send_json({
                                    "type": "auth_success",
                                    "user_id": user_id
                                })
                        except:
                            await websocket.send_json({
                                "type": "auth_error",
                                "message": "Invalid token"
                            })
                
                # Handle ping/pong for keepalive
                elif message.get("type") == "ping":
                    await websocket.send_json({"type": "pong"})
                    
            except WebSocketDisconnect:
                break
            except Exception as e:
                logger.error(f"WebSocket error: {e}")
                break
                
    except Exception as e:
        logger.error(f"WebSocket connection error: {e}")
    finally:
        ws_manager.disconnect(websocket, user_id)

# Helper function to broadcast snag updates
async def broadcast_snag_update(event_type: str, snag_data: dict):
    """Broadcast snag updates to all connected clients"""
    await ws_manager.broadcast({
        "type": "snag_update",
        "event": event_type,
        "data": snag_data,
        "timestamp": datetime.utcnow().isoformat()
    })

# Helper function to broadcast notification
async def broadcast_notification(user_id: str, notification: dict):
    """Send notification to specific user via WebSocket"""
    await ws_manager.send_to_user(user_id, {
        "type": "notification",
        "data": notification,
        "timestamp": datetime.utcnow().isoformat()
    })

# ==================== Initialize Default Manager ====================

@app.on_event("startup")
async def create_default_manager():
    # Check if any manager exists
    manager = await db.users.find_one({"role": UserRole.MANAGER})
    if not manager:
        # Create default manager
        default_manager = {
            "email": "manager@pmc.com",
            "password": get_password_hash("manager123"),
            "name": "Default Manager",
            "role": UserRole.MANAGER,
            "phone": None,
            "push_token": None,
            "created_at": datetime.utcnow()
        }
        await db.users.insert_one(default_manager)
        logger.info("Default manager created: manager@pmc.com / manager123")

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
